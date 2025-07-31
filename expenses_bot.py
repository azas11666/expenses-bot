import os
import whisper
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    MessageHandler,
    CommandHandler,
    ContextTypes,
    filters
)

TOKEN = os.getenv("BOT_TOKEN")

model = whisper.load_model("base")

EXCEL_FILE = "expenses.xlsx"
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["التاريخ", "الصنف", "المبلغ"])
    wb.save(EXCEL_FILE)

logging.basicConfig(level=logging.INFO)

def extract_expenses(text):
    import re
    pattern = r'(\d+)\s*ريال(?:.*?على)?\s*([\u0600-\u06FF]+)'
    matches = re.findall(pattern, text)
    return [(category.strip(), int(amount)) for amount, category in matches]

def save_to_excel(expenses):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for category, amount in expenses:
        ws.append([datetime.now().strftime("%Y-%m-%d"), category, amount])
    wb.save(EXCEL_FILE)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 مرحباً بك! أرسل رسالة صوتية تحتوي على مصروفاتك، وسأقوم بتحويلها وتسجيلها في ملف Excel."
    )

async def voice_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    voice = update.message.voice
    file = await context.bot.get_file(voice.file_id)
    file_path = "voice.ogg"
    await file.download_to_drive(file_path)

    wav_path = "voice.wav"
    os.system(f"ffmpeg -i {file_path} -ar 16000 -ac 1 {wav_path} -y")

    result = model.transcribe(wav_path, language="ar")
    text = result["text"]

    await update.message.reply_text(f"📄 النص المستخرج:\n{text}")

    expenses = extract_expenses(text)
    if expenses:
        save_to_excel(expenses)
        await update.message.reply_text("✅ تم تسجيل المصروفات في ملف Excel.")
    else:
        await update.message.reply_text("❌ لم أتمكن من استخراج مصروفات من الرسالة.")

if __name__ == "__main__":
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.VOICE, voice_handler))
    app.run_polling()
